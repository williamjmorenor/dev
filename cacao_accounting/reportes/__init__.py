# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2025 - 2026 William José Moreno Reyes

"""Reportes operativos de subledgers, aging, Kardex y reconciliaciones."""

from __future__ import annotations

import csv
from dataclasses import replace
from datetime import date
from decimal import Decimal
from decimal import DecimalException
from io import BytesIO, StringIO
from typing import Any

from flask import Blueprint, flash, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from cacao_accounting.database import Entity, UserFormPreference, database
from cacao_accounting.decorators import modulo_activo, verifica_acceso
from cacao_accounting.form_preferences import get_form_preference, reset_form_preference, save_form_preference
from cacao_accounting.reportes.services import (
    AgingFilters,
    FinancialReportFilters,
    KardexFilters,
    OperationalReportFilters,
    SubledgerFilters,
    get_account_movement_detail,
    get_aging_report,
    get_ar_ap_subledger,
    get_batch_report,
    get_balance_sheet_report,
    get_gross_margin,
    get_income_statement_report,
    get_inventory_valuation,
    get_kardex,
    get_purchases_by_item,
    get_purchases_by_supplier,
    get_reconciliation_report,
    get_sales_by_customer,
    get_sales_by_item,
    get_serial_report,
    get_stock_balance,
    get_trial_balance_report,
)
from cacao_accounting.version import APPNAME

try:  # pragma: no cover - fallback defensivo para contextos sin Flask-Babel inicializado.
    from flask_babel import gettext as _babel_gettext
except ImportError:  # pragma: no cover

    def _(value: str) -> str:
        return value

else:

    def _(value: str) -> str:
        try:
            return _babel_gettext(value)
        except (KeyError, RuntimeError):
            return value


reportes = Blueprint("reportes", __name__, template_folder="templates")

_COLUMN_LABELS = {
    "posting_date": "Posting Date",
    "accounting_period": "Period",
    "document_no": "Voucher",
    "voucher_type": "Type",
    "account_code": "Account",
    "account_name": "Account Name",
    "debit": "Debit",
    "credit": "Credit",
    "running_balance": "Final Balance",
    "currency": "Currency",
    "ledger": "Ledger",
    "company": "Company",
    "opening_balance": "Opening Balance",
    "ending_balance": "Final Balance",
    "cost_center": "Cost Center",
    "unit": "Unit",
    "project": "Project",
    "party_type": "Party Type",
    "party_id": "Party",
    "created_by": "User",
    "created_at": "Creation Date",
    "line_comment": "Reference",
    "voucher_status": "Status",
    "section": "Section",
    "amount": "Amount",
}
_MONEY_COLUMNS = {
    "debit",
    "credit",
    "difference",
    "opening_balance",
    "ending_balance",
    "running_balance",
    "amount",
    "assets",
    "liabilities",
    "equity",
    "period_profit",
    "income",
    "cost",
    "expense",
    "gross_profit",
    "net_profit",
}
_RIGHT_ALIGN_COLUMNS = _MONEY_COLUMNS | {"level"}
_ALWAYS_VISIBLE_COLUMNS = {"debit", "credit", "difference", "account_code", "account_name", "section", "amount"}
_EMPTY_CELL_VALUE = "—"
_FINANCIAL_FILTER_FIELDS = (
    "company",
    "ledger",
    "accounting_period",
    "voucher_number",
    "account_code",
    "account_from",
    "account_to",
    "cost_center_code",
    "unit_code",
    "project_code",
    "party_type",
    "party_id",
    "voucher_type",
    "status",
    "include_running_balance",
    "page_size",
    "sort_by",
    "sort_dir",
    "group_by",
)


def _format_number(value: object) -> str:
    try:
        amount = value if isinstance(value, Decimal) else Decimal(str(value))
    except DecimalException:
        return _EMPTY_CELL_VALUE
    formatted = f"{abs(amount):,.2f}"
    return f"({formatted})" if amount < 0 else formatted


def _column_label(column: str, ledger_currency: str | None) -> str:
    label = _(_COLUMN_LABELS.get(column, column.replace("_", " ").title()))
    if column in _MONEY_COLUMNS and ledger_currency:
        return f"{label} ({ledger_currency})"
    return label


def _format_cell(column: str, value: object, ledger_currency: str | None) -> str:
    if value is None or value == "":
        return _EMPTY_CELL_VALUE
    if column in _MONEY_COLUMNS:
        return _format_number(value)
    if column == "posting_date" and isinstance(value, date):
        return value.isoformat()
    if column == "voucher_status":
        return _("Cancelado") if str(value).lower() == "cancelled" else _("Contabilizado")
    if column == "section":
        section_labels = {
            "assets": _("ACTIVOS"),
            "liabilities": _("PASIVOS"),
            "equity": _("PATRIMONIO"),
            "income": _("INGRESOS"),
            "cost": _("COSTOS"),
            "expense": _("GASTOS"),
            "gross_profit": _("UTILIDAD BRUTA"),
            "net_profit": _("UTILIDAD NETA"),
        }
        return section_labels.get(str(value), str(value))
    return str(value)


def _build_context_summary(report, report_filters: FinancialReportFilters) -> dict[str, str]:
    ledger_label = report_filters.ledger or "—"
    if report_filters.ledger and report.ledger_currency:
        ledger_label = f"{report_filters.ledger} ({report.ledger_currency})"
    status_value = report_filters.status or "submitted"
    status_label = _("Cancelado") if status_value == "cancelled" else _("Contabilizado")
    return {
        "company": report_filters.company,
        "ledger": ledger_label,
        "period": report_filters.accounting_period or "—",
        "status": status_label,
        "records": str(report.total_rows),
    }


def _is_report_balanced(raw_totals: dict[str, Decimal]) -> bool:
    difference = raw_totals.get("difference")
    if difference is None:
        return False
    try:
        return Decimal(str(difference)) == Decimal("0")
    except DecimalException:
        return False


def _report_form_key(report_code: str) -> str:
    return f"reports.financial.{report_code}"


def _load_report_view_options(report_code: str) -> list[str]:
    preferences = (
        database.session.execute(
            database.select(UserFormPreference.view_key)
            .filter_by(user_id=str(current_user.id), form_key=_report_form_key(report_code))
            .order_by(UserFormPreference.view_key.asc())
        )
        .scalars()
        .all()
    )
    views = ["default"]
    views.extend([view for view in preferences if view != "default"])
    return views


def _extract_filter_payload() -> dict[str, str]:
    payload: dict[str, str] = {}
    for key in _FINANCIAL_FILTER_FIELDS:
        value = request.args.get(key)
        if value:
            payload[key] = value
    visible_columns = request.args.getlist("visible_columns")
    if visible_columns:
        payload["visible_columns"] = ",".join(visible_columns)
    return payload


def _restore_filters_from_view(filters: FinancialReportFilters, report_code: str, view_key: str) -> FinancialReportFilters:
    preference = get_form_preference(str(current_user.id), _report_form_key(report_code), view_key)
    payload = preference.get("filters", {})
    if not isinstance(payload, dict):
        return filters
    try:
        page_size = max(int(payload.get("page_size") or filters.page_size), 1)
    except (TypeError, ValueError):
        page_size = max(filters.page_size, 1)
    return replace(
        filters,
        company=str(payload.get("company") or filters.company),
        ledger=str(payload.get("ledger") or "") or None,
        accounting_period=str(payload.get("accounting_period") or "") or None,
        voucher_number=str(payload.get("voucher_number") or "") or None,
        account_code=str(payload.get("account_code") or "") or None,
        account_from=str(payload.get("account_from") or "") or None,
        account_to=str(payload.get("account_to") or "") or None,
        cost_center_code=str(payload.get("cost_center_code") or "") or None,
        unit_code=str(payload.get("unit_code") or "") or None,
        project_code=str(payload.get("project_code") or "") or None,
        party_type=str(payload.get("party_type") or "") or None,
        party_id=str(payload.get("party_id") or "") or None,
        voucher_type=str(payload.get("voucher_type") or "") or None,
        status=str(payload.get("status") or filters.status or "submitted"),
        include_running_balance=str(payload.get("include_running_balance") or "").lower() in {"1", "true", "yes", "on"},
        page_size=page_size,
        sort_by=str(payload.get("sort_by") or filters.sort_by),
        sort_dir=str(payload.get("sort_dir") or filters.sort_dir),
        page=1,
    )


def _handle_saved_view_action(report_code: str, filters: FinancialReportFilters) -> tuple[FinancialReportFilters, str]:
    view_key = (request.args.get("saved_view") or "default").strip() or "default"
    action = request.args.get("view_action")
    if action == "save" and view_key:
        payload = {
            "schema_version": 1,
            "filters": _extract_filter_payload(),
            "columns": [
                {
                    "field": column,
                    "label": column,
                    "visible": True,
                    "width": 1,
                    "required": False,
                }
                for column in request.args.getlist("visible_columns")
            ],
        }
        save_form_preference(
            user_id=str(current_user.id),
            form_key=_report_form_key(report_code),
            view_key=view_key,
            payload=payload,
        )
        flash(_("Vista guardada correctamente."), "success")
    elif action == "reset" and view_key != "default":
        reset_form_preference(str(current_user.id), _report_form_key(report_code), view_key)
        flash(_("Vista eliminada correctamente."), "warning")
        view_key = "default"
    elif action == "apply" and view_key != "default":
        filters = _restore_filters_from_view(filters, report_code, view_key)
    return filters, view_key


def _resolve_view_context(report_code: str, filters: FinancialReportFilters) -> tuple[FinancialReportFilters, str, list[str]]:
    resolved_filters, selected_view = _handle_saved_view_action(report_code, filters)
    return resolved_filters, selected_view, _load_report_view_options(report_code)


def _preferred_columns_from_view(report_code: str, view_key: str) -> list[str]:
    if view_key == "default":
        return []
    preference = get_form_preference(str(current_user.id), _report_form_key(report_code), view_key)
    columns = preference.get("columns", [])
    if not isinstance(columns, list):
        return []
    visible: list[str] = []
    for column in columns:
        if not isinstance(column, dict):
            continue
        if not bool(column.get("visible", True)):
            continue
        field = str(column.get("field") or "").strip()
        if field:
            visible.append(field)
    return visible


def _preferred_group_by_from_view(report_code: str, view_key: str) -> str:
    if view_key == "default":
        return ""
    preference = get_form_preference(str(current_user.id), _report_form_key(report_code), view_key)
    filters = preference.get("filters", {})
    if not isinstance(filters, dict):
        return ""
    return str(filters.get("group_by") or "")


def _resolve_company(company_code: str) -> str:
    requested_company = company_code or "cacao"
    company_exists = database.session.execute(
        database.select(Entity.code).where(Entity.code == requested_company)
    ).scalar_one_or_none()
    if company_exists is not None:
        return requested_company
    default_company = database.session.execute(
        database.select(Entity.code).order_by(Entity.default.desc(), Entity.code.asc())
    ).scalar_one_or_none()
    return default_company or "cacao"


def _build_drill_down_url(values: dict[str, object], company: str, ledger: str | None) -> str | None:
    account_code = values.get("account_code")
    if account_code in (None, "", _EMPTY_CELL_VALUE):
        return None
    query: dict[str, Any] = {
        "company": company,
        "account_code": str(account_code),
    }
    if ledger:
        query["ledger"] = ledger
    return url_for("reportes.account_movement", **query)


def _build_voucher_url(values: dict[str, object]) -> str | None:
    voucher_type = str(values.get("voucher_type") or "").lower()
    voucher_id = str(values.get("document_no") or values.get("voucher_id") or "").strip()
    if voucher_type in {"journal_entry", "comprobante_contable"} and voucher_id:
        return url_for("contabilidad.ver_comprobante", identifier=voucher_id)
    return None


def _date_arg(name: str) -> date | None:
    value = request.args.get(name)
    return date.fromisoformat(value) if value else None


def _int_arg(name: str, default: int) -> int:
    value = request.args.get(name, default=str(default))
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _bool_arg(name: str) -> bool:
    return request.args.get(name, "").lower() in {"1", "true", "yes", "on"}


def _financial_filters() -> FinancialReportFilters:
    company_code = _resolve_company(request.args.get("company", "cacao"))
    return FinancialReportFilters(
        company=company_code,
        ledger=request.args.get("ledger") or None,
        accounting_period=request.args.get("accounting_period") or None,
        voucher_number=request.args.get("voucher_number") or None,
        account_code=request.args.get("account_code") or None,
        account_from=request.args.get("account_from") or None,
        account_to=request.args.get("account_to") or None,
        cost_center_code=request.args.get("cost_center_code") or None,
        unit_code=request.args.get("unit_code") or None,
        project_code=request.args.get("project_code") or None,
        party_type=request.args.get("party_type") or None,
        party_id=request.args.get("party_id") or None,
        voucher_type=request.args.get("voucher_type") or None,
        status=request.args.get("status") or "submitted",
        include_running_balance=_bool_arg("include_running_balance"),
        page=max(_int_arg("page", 1), 1),
        page_size=max(_int_arg("page_size", 100), 1),
        sort_by=request.args.get("sort_by", "posting_date"),
        sort_dir=request.args.get("sort_dir", "asc"),
        export_all=False,
    )


def _report_to_matrix(report) -> tuple[list[str], list[list[object]]]:
    columns = report.columns or (list(report.rows[0].values.keys()) if report.rows else [])
    data_rows = [[row.values.get(column) for column in columns] for row in report.rows]
    return columns, data_rows


def _export_financial_report(report, report_code: str, title: str, report_filters: FinancialReportFilters):
    export_format = request.args.get("export")
    if export_format not in {"csv", "xlsx"}:
        return None

    columns, rows = _report_to_matrix(report)
    if export_format == "csv":
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(columns)
        writer.writerows(rows)
        return send_file(
            BytesIO(buffer.getvalue().encode("utf-8")),
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"{report_code}.csv",
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = report_code[:31]
    sheet.append([title])
    sheet.append([_("Fecha de generación"), date.today().isoformat()])
    sheet.append([_("Usuario"), getattr(current_user, "user", "")])
    sheet.append([])
    if columns:
        localized_headers = [_column_label(column, report.ledger_currency) for column in columns]
        sheet.append(localized_headers)
        sheet.freeze_panes = "A5"
    for row in rows:
        sheet.append([_format_cell(column, row[index], report.ledger_currency) for index, column in enumerate(columns)])
    if report.totals:
        sheet.append([])
    for total_name, total_value in report.totals.items():
        sheet.append(
            [
                _("TOTAL"),
                _column_label(total_name, report.ledger_currency),
                _format_cell(total_name, total_value, report.ledger_currency),
            ]
        )
    for column_cells in sheet.columns:
        values = [str(cell.value or "") for cell in column_cells]
        max_length = max((len(value) for value in values), default=10)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 60)
    for column in range(1, sheet.max_column + 1):
        sheet.cell(row=4, column=column).alignment = Alignment(horizontal="center")

    filters_sheet = workbook.create_sheet(_("Filtros"))
    filters_sheet.append([_("Filtro"), _("Valor")])
    for key in _FINANCIAL_FILTER_FIELDS:
        value = getattr(report_filters, key, None)
        if value in (None, "", False):
            continue
        filters_sheet.append([_(key.replace("_", " ").title()), str(value)])
    filters_sheet.freeze_panes = "A2"
    content = BytesIO()
    workbook.save(content)
    content.seek(0)
    return send_file(
        content,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{report_code}.xlsx",
    )


def _render_financial_report(
    report_code: str,
    report_title: str,
    report,
    report_filters: FinancialReportFilters,
    saved_view: str,
    saved_views: list[str],
):
    export_response = _export_financial_report(report, report_code, report_title, report_filters)
    if export_response is not None:
        return export_response
    selected_columns = request.args.getlist("visible_columns")
    if not selected_columns:
        selected_columns = _preferred_columns_from_view(report_code, saved_view)
    columns = report.columns or []
    if selected_columns:
        columns = [column for column in columns if column in selected_columns]
    display_columns = [
        column
        for column in columns
        if any((row.values.get(column) not in (None, "", "—") for row in report.rows)) or column in _ALWAYS_VISIBLE_COLUMNS
    ]
    if not display_columns:
        display_columns = columns
    display_headers = {column: _column_label(column, report.ledger_currency) for column in display_columns}
    row_metadata = []
    child_counts: dict[str, int] = {}
    for row in report.rows:
        account_code = str(row.values.get("account_code") or "")
        if not account_code:
            continue
        parent_code = ".".join(account_code.split(".")[:-1])
        if parent_code:
            child_counts[parent_code] = child_counts.get(parent_code, 0) + 1
    for row in report.rows:
        account_code = str(row.values.get("account_code") or "")
        parent_code = ".".join(account_code.split(".")[:-1]) if account_code else ""
        row_metadata.append(
            {
                "code": account_code,
                "parent": parent_code,
                "has_children": bool(child_counts.get(account_code)),
                "level": int(row.values.get("level") or account_code.count(".") + 1 if account_code else 0),
                "drilldown_url": _build_drill_down_url(row.values, report_filters.company, report_filters.ledger),
                "voucher_url": _build_voucher_url(row.values),
            }
        )
    display_rows: list[dict[str, object]] = []
    for index, row in enumerate(report.rows):
        formatted_row: dict[str, object] = {
            column: _format_cell(column, row.values.get(column), report.ledger_currency) for column in display_columns
        }
        formatted_row["__meta"] = row_metadata[index]
        display_rows.append(formatted_row)
    group_by = request.args.get("group_by") or _preferred_group_by_from_view(report_code, saved_view)
    grouped_rows: list[dict[str, object]] = []
    if report_code == "account-movement" and group_by and group_by in display_columns:
        current_group = None
        for row in display_rows:
            group_value = row.get(group_by, _EMPTY_CELL_VALUE)
            if group_value != current_group:
                group_row: dict[str, object] = {
                    "__row_type": "group",
                    "__group_title": f"{_(group_by.replace('_', ' ').title())}: {group_value}",
                }
                grouped_rows.append(group_row)
                current_group = group_value
            grouped_rows.append(row)
    else:
        grouped_rows = display_rows
    display_totals = {key: _format_cell(key, value, report.ledger_currency) for key, value in report.totals.items()}
    return render_template(
        "reportes/financial_report.html",
        titulo=f"{report_title} - {APPNAME}",
        report_code=report_code,
        report_title=report_title,
        rows=report.rows,
        columns=display_columns,
        display_headers=display_headers,
        display_rows=grouped_rows,
        totals=display_totals,
        total_rows=report.total_rows,
        page=report.page,
        page_size=report.page_size,
        ledger_currency=report.ledger_currency,
        context_summary=_build_context_summary(report, report_filters),
        right_align_columns=_RIGHT_ALIGN_COLUMNS,
        is_balanced=_is_report_balanced(report.totals),
        saved_view=saved_view,
        saved_views=saved_views,
        selected_columns=display_columns,
        all_columns=report.columns or [],
        group_by=group_by,
    )


@reportes.route("/reports/account-movement")
@login_required
@modulo_activo("accounting")
@verifica_acceso("accounting")
def account_movement():
    """Reporte unificado de detalle de movimiento contable."""
    filters, selected_view, saved_views = _resolve_view_context("account-movement", _financial_filters())
    report = get_account_movement_detail(filters)
    if request.args.get("export") in {"csv", "xlsx"}:
        export_report = get_account_movement_detail(replace(filters, export_all=True, page=1))
        return _render_financial_report(
            "account-movement",
            _("Detalle de Movimiento Contable"),
            export_report,
            filters,
            selected_view,
            saved_views,
        )
    return _render_financial_report(
        "account-movement",
        _("Detalle de Movimiento Contable"),
        report,
        filters,
        selected_view,
        saved_views,
    )


@reportes.route("/reports/trial-balance")
@login_required
@modulo_activo("accounting")
@verifica_acceso("accounting")
def trial_balance():
    """Reporte de balanza de comprobación."""
    filters, selected_view, saved_views = _resolve_view_context("trial-balance", _financial_filters())
    report = get_trial_balance_report(filters)
    return _render_financial_report(
        "trial-balance",
        _("Balanza de Comprobación"),
        report,
        filters,
        selected_view,
        saved_views,
    )


@reportes.route("/reports/income-statement")
@login_required
@modulo_activo("accounting")
@verifica_acceso("accounting")
def income_statement():
    """Reporte de estado de resultado."""
    filters, selected_view, saved_views = _resolve_view_context("income-statement", _financial_filters())
    report = get_income_statement_report(filters)
    return _render_financial_report(
        "income-statement",
        _("Estado de Resultado"),
        report,
        filters,
        selected_view,
        saved_views,
    )


@reportes.route("/reports/balance-sheet")
@login_required
@modulo_activo("accounting")
@verifica_acceso("accounting")
def balance_sheet():
    """Reporte de balance general."""
    filters, selected_view, saved_views = _resolve_view_context("balance-sheet", _financial_filters())
    report = get_balance_sheet_report(filters)
    return _render_financial_report(
        "balance-sheet",
        _("Balance General"),
        report,
        filters,
        selected_view,
        saved_views,
    )


@reportes.route("/reports/subledger")
@login_required
@modulo_activo("accounting")
def subledger():
    """Reporte AR/AP por documento."""
    company = request.args.get("company", "cacao")
    party_type = request.args.get("party_type", "customer")
    report = get_ar_ap_subledger(
        SubledgerFilters(
            company=company,
            party_type=party_type,
            party_id=request.args.get("party_id") or None,
            as_of_date=_date_arg("as_of_date"),
        )
    )
    return render_template(
        "reportes/report_table.html",
        titulo="Subledger AR/AP - " + APPNAME,
        report_title="Subledger AR/AP",
        rows=report.rows,
        totals=report.totals,
    )


@reportes.route("/reports/aging")
@login_required
@modulo_activo("accounting")
def aging():
    """Reporte aging AR/AP."""
    company = request.args.get("company", "cacao")
    party_type = request.args.get("party_type", "customer")
    report = get_aging_report(
        AgingFilters(
            company=company,
            party_type=party_type,
            party_id=request.args.get("party_id") or None,
            as_of_date=_date_arg("as_of_date") or date.today(),
        )
    )
    return render_template(
        "reportes/report_table.html",
        titulo="Aging AR/AP - " + APPNAME,
        report_title="Aging AR/AP",
        rows=report.rows,
        totals=report.totals,
    )


@reportes.route("/reports/kardex")
@login_required
@modulo_activo("inventory")
def kardex():
    """Reporte Kardex."""
    report = get_kardex(
        KardexFilters(
            company=request.args.get("company", "cacao"),
            item_code=request.args.get("item_code") or None,
            warehouse=request.args.get("warehouse") or None,
            date_from=_date_arg("date_from"),
            date_to=_date_arg("date_to"),
        )
    )
    return render_template(
        "reportes/report_table.html",
        titulo="Kardex - " + APPNAME,
        report_title="Kardex",
        rows=report.rows,
        totals=report.totals,
    )


@reportes.route("/reports/reconciliations")
@login_required
@modulo_activo("accounting")
def reconciliations():
    """Reporte de reconciliaciones."""
    report = get_reconciliation_report(
        company=request.args.get("company", "cacao"),
        as_of_date=_date_arg("as_of_date"),
    )
    return render_template(
        "reportes/report_table.html",
        titulo="Reconciliaciones - " + APPNAME,
        report_title="Reconciliaciones",
        rows=report.rows,
        totals=report.totals,
    )


def _operational_filters() -> OperationalReportFilters:
    return OperationalReportFilters(
        company=request.args.get("company", "cacao"),
        date_from=_date_arg("date_from"),
        date_to=_date_arg("date_to"),
        party_id=request.args.get("party_id") or None,
        item_code=request.args.get("item_code") or None,
        warehouse=request.args.get("warehouse") or None,
    )


def _render_operational_report(report_name: str, report):
    return render_template(
        "reportes/report_table.html",
        titulo=report_name + " - " + APPNAME,
        report_title=report_name,
        rows=report.rows,
        totals=report.totals,
    )


@reportes.route("/reports/purchases-by-supplier")
@login_required
@modulo_activo("purchases")
def purchases_by_supplier():
    """Genera reporte de compras agrupadas por proveedor."""
    return _render_operational_report("Compras por Proveedor", get_purchases_by_supplier(_operational_filters()))


@reportes.route("/reports/purchases-by-item")
@login_required
@modulo_activo("purchases")
def purchases_by_item():
    """Genera reporte de compras agrupadas por articulo."""
    return _render_operational_report("Compras por Item", get_purchases_by_item(_operational_filters()))


@reportes.route("/reports/sales-by-customer")
@login_required
@modulo_activo("sales")
def sales_by_customer():
    """Genera reporte de ventas agrupadas por cliente."""
    return _render_operational_report("Ventas por Cliente", get_sales_by_customer(_operational_filters()))


@reportes.route("/reports/sales-by-item")
@login_required
@modulo_activo("sales")
def sales_by_item():
    """Genera reporte de ventas agrupadas por articulo."""
    return _render_operational_report("Ventas por Item", get_sales_by_item(_operational_filters()))


@reportes.route("/reports/gross-margin")
@login_required
@modulo_activo("sales")
def gross_margin():
    """Genera reporte de margen bruto por ventas."""
    return _render_operational_report("Margen Bruto", get_gross_margin(_operational_filters()))


@reportes.route("/reports/stock-balance")
@login_required
@modulo_activo("inventory")
def stock_balance():
    """Genera reporte de balance de stock por articulo y bodega."""
    return _render_operational_report("Stock Balance", get_stock_balance(_operational_filters()))


@reportes.route("/reports/inventory-valuation")
@login_required
@modulo_activo("inventory")
def inventory_valuation():
    """Genera reporte de valoracion del inventario."""
    return _render_operational_report("Valoracion de Inventario", get_inventory_valuation(_operational_filters()))


@reportes.route("/reports/batches")
@login_required
@modulo_activo("inventory")
def batches():
    """Genera reporte de lotes de inventario."""
    return _render_operational_report("Lotes", get_batch_report(_operational_filters()))


@reportes.route("/reports/serials")
@login_required
@modulo_activo("inventory")
def serials():
    """Genera reporte de numeros de serie de inventario."""
    return _render_operational_report("Seriales", get_serial_report(_operational_filters()))
